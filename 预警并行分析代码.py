import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill # 颜色
from openpyxl.utils import get_column_letter # 给每页加日期

# 今日
target_date = datetime.datetime(2024,2,25).strftime("%Y-%m-%d")
# 昨日
# target_date_y = (datetime.datetime(2024,2,25)-datetime.timedelta(days=1)).strftime("%Y-%m-%d")

output_path = "/Users/wenxinjiang/Desktop/预警并行结果分析" + target_date + ".xlsx"
input_path = "/Users/wenxinjiang/Desktop/德勤/新老模型得分对比/"



# Extract data on target_date
# 1. 系统 主体预警得分，等级
sys_data = pd.read_pickle(input_path+"entity_res_t.pkl")
sys_df = pd.DataFrame(sys_data)

# 2. 新模型 主体预警得分，等级
new_data = pd.read_pickle(input_path+"warn_entityscore.pkl")
# new_df = pd.DataFrame(new_data)
new_df = new_data[new_data['日期'] == target_date] # 现在新模型数据已经开始并行了，需要筛选日期了

# 3. 新模型 主体input coef，传导关系
new_input_data = pd.read_pickle(input_path+"rela_pd_level.pkl")
new_input_df = pd.DataFrame(new_input_data)
# 4. 过渡模型 主体预警得分，等级
trans_data = pd.read_pickle(input_path+'trans_amount.pkl')
trans_df = pd.DataFrame(trans_data)
trans_df = trans_df[['主体名称','旧模型新分数','预警等级_trans','预警得分_trans']]

# 5. 滑档区间
# intervals_final_data = pd.read_pickle(input_path+'aintervals_final.pkl')
# intervals_final_df = pd.DataFrame(intervals_final_data)
all_interval_data = pd.read_pickle(input_path+'all_intervals.pkl')
all_intervals_df = pd.DataFrame(all_interval_data)

# 系统只会包含需要分析那日当天的数据，所以不用筛选日期？？
# sys_df = sys_data[sys_data['日期'] == target_date]
# new_input_df = new_input_data[new_input_data['日期'] == target_date]

# # 4. 系统 过去三十天的记录
# sys_3_data = pd.read_pickle(input_path+"sys_3_data.pkl")
# sys_past30_df = pd.DataFrame(sys_3_data)
# # 5. 新模型 过去三十天的记录
# new02 = pd.read_pickle(input_path+"new02.pkl")
# new_past30_df = pd.DataFrame(new02)

# -------------------------Excel Output------------------------------------
#
# -------------------------1--------------------------------------------
# Sheet 1: 新老模型分数等级明细
# columns：主体名称	风险等级_sys	风险等级_new	风险等级_trans
#           预警等级_sys 预警等级_new 预警等级_trans
#           预警得分_sys	预警得分_new 预警得分_trans
new_df = new_df.copy()
# 首先，我们将三个dataframes合并成一个，以主体名称和time为键，这样我们可以在一个dataframe中看到每个entity在不同数据集中的分数
# 转换数据类型 to_numeric
sys_df['预警等级'] = pd.to_numeric(sys_df['预警等级'], errors='coerce')
# new_df['预警等级'] = pd.to_numeric(new_df['预警等级'], errors='coerce')
new_df.loc[:, '预警等级'] = pd.to_numeric(new_df['预警等级'], errors='coerce')
trans_df['预警等级_trans'] = pd.to_numeric(trans_df['预警等级_trans'], errors='coerce')
# 预警等级映射到风险等级
risk_level_mapping = {
    10: '高', 9: '中高', 8: '中高', 7: '中', 6: '中',
    5: '低', 4: '低', 3: '正常', 2: '正常', 1: '正常', 0: '正常'
}
risk_levels = ['高', '中高','中', '低', '正常']
# 映射预警等级到风险等级
sys_df['风险等级'] = sys_df['预警等级'].map(risk_level_mapping)
# new_df['风险等级'] = new_df['预警等级'].map(risk_level_mapping)
new_df.loc[:, '风险等级'] = new_df['预警等级'].map(risk_level_mapping)
trans_df['风险等级_trans'] = trans_df['预警等级_trans'].map(risk_level_mapping)

# 使用merge函数进行合并，并给每个分数列一个唯一的名字
merged = sys_df.merge(new_df, on=['主体名称'], how='outer', suffixes=('_sys', '_new'))
merged_df = merged.merge(trans_df, on=['主体名称'], how='outer')

# 调整列的顺序
column_order = ['主体名称',
                '风险等级_sys','风险等级_new', '风险等级_trans',
                '预警等级_sys', '预警等级_new', '预警等级_trans',
                '预警得分_sys', '旧模型新分数', '预警得分_new', '预警得分_trans']
merged_df = merged_df[column_order]

# 将合并后的数据写入Excel文件的一个新工作表中
merged_df.to_excel(output_path, sheet_name='新老模型分数等级明细', index=False)

# -------------------------2--------------------------------------------
# Sheet 2
# 新模型指标明细
column_new_df_coef = ['主体名称','pd','daily_level','hist_level','final_level',
                      'daily_level_rela','hist_level_rela','final_level_rela',
                      # 'effect','pd_score_relation','pd_score_relation_mod',
                'score_news', 'score_price', 'score_issr', 'score_basic']
new_df_coef = new_df.merge(new_input_df, on=['主体名称'], how='outer')
new_df_coef = new_df_coef[column_new_df_coef]
new_df_coef = new_df_coef.rename(columns = {'daily_level_rela':'daily_level主体关系传导后',
                                            'hist_level_rela':'hist_level主体关系传导后',
                                            'final_level_rela':'final_level主体关系传导后'})
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
    new_df_coef.to_excel(writer, sheet_name='新模型指标明细', index=False)

# -------------------------滑档区间--------------------------------------------
# Sheet ？：滑档区间
# Sheet ？：新模型滑档区间

# Identify numeric columns for conversion
columns_to_convert = all_intervals_df.columns # Exclude '预警等级' as it seems categorical

# Initialize an empty DataFrame for the converted intervals
df_converted = pd.DataFrame()

# Convert each column
for column in columns_to_convert:
    min_values = all_intervals_df[column]
    max_values = all_intervals_df[column].shift(-1).fillna(
        1)  # Use the next row's value as max, fill the last with its own value

    # Format intervals as [min, max) and round to 3 decimal places
    df_converted[column] = ['[' + format(min_value, '.3f') + ', ' + format(max_value, '.4f') + ')' for
                            min_value, max_value in zip(min_values, max_values)]

df_converted = df_converted.iloc[:-1, :]

with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
    df_converted.to_excel(writer, sheet_name='滑档区间')
    # intervals_final_df.to_excel(writer, sheet_name='新模型滑档区间')


# -------------------------3--------------------------------------------
# Sheet 3：预警等级分布
# 创建等级列表，从0到10
full_levels = [10,9,8,7,6,5,4,3,2,1,0]
level_summary_df = pd.DataFrame({'等级': full_levels})

# 对每个数据集计算每个预警等级的entity数量
level_summary_df['预警等级_sys'] = level_summary_df['等级'].map(sys_df['预警等级'].value_counts()).fillna(0).astype(int)
level_summary_df['预警等级_new'] = level_summary_df['等级'].map(new_df['预警等级'].value_counts()).fillna(0).astype(int)
level_summary_df['预警等级_trans'] = level_summary_df['等级'].map(trans_df['预警等级_trans'].value_counts()).fillna(0).astype(int)

# 将统计数据写入Excel文件的一个新工作表中
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
    level_summary_df.to_excel(writer, sheet_name='预警等级分布', index=False)


# -------------------------4--------------------------------------------
# Sheet 4：风险等级分布

# 计算每个风险等级的数量
risk_level_counts_sys = sys_df['风险等级'].value_counts().reindex(risk_levels, fill_value='null')
risk_level_counts_new = new_df['风险等级'].value_counts().reindex(risk_levels, fill_value='null')
risk_level_counts_trans = trans_df['风险等级_trans'].value_counts().reindex(risk_levels, fill_value='null')

# 创建一个新的DataFrame以存储计数
risk_level_summary = pd.DataFrame({
    '风险等级': risk_levels,
    'sys': risk_level_counts_sys.values,
    'new': risk_level_counts_new.values,
    'trans': risk_level_counts_trans.values
})

# 写入Excel文件
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
    risk_level_summary.to_excel(writer, sheet_name='风险等级分布', index=False)


# -------------------------5--------------------------------------------
# Sheet 5：风险变化主体清单
# 合并数据集，对比风险等级
merged_df = sys_df.merge(new_df, on='主体名称', suffixes=('_sys', '_new'))
diff_df = merged_df[merged_df['风险等级_sys'] != merged_df['风险等级_new']]

# 结果DataFrame将只包含风险等级有差异的主体
risk_level_diff_df = diff_df[['主体名称', '风险等级_sys', '风险等级_new']]

# 写入Excel文件
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
    risk_level_diff_df.to_excel(writer, sheet_name='风险变化主体清单', index=False)


# -------------------------6--------------------------------------------
# Sheet 6 风险变化主体清单透视表
# 使用之前的风险变化主体清单risk_level_diff_df来创建透视表
risk_pivot_table_df = pd.pivot_table(
    risk_level_diff_df,
    index='风险等级_sys',
    columns='风险等级_new',
    values='主体名称',
    aggfunc='count',
    # fill_value=0,
    margins=True,  # 添加总计
    margins_name='总计'  # 总计的名称
)

# 确保索引和列的顺序符合预定的风险等级顺序，同时添加总计
risk_pivot_table_df = risk_pivot_table_df.reindex(['风险等级（旧）']+risk_levels + ['总计'])
risk_pivot_table_df = risk_pivot_table_df.reindex(columns=['风险等级（新）']+risk_levels + ['总计'])

# 为行和列索引添加说明
risk_pivot_table_df.index.name = 'Count of Entities'

# 写入Excel文件
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
    risk_pivot_table_df.to_excel(writer, sheet_name='风险等级变化主体透视表')




# -------------------------7--------------------------------------------
# Sheet 7: Diff Sys-New等级
diff_sys_new_df = merged_df

# 计算差异
diff_sys_new_df['预警等级_sys'] = pd.to_numeric(diff_sys_new_df['预警等级_sys'], errors='coerce')
diff_sys_new_df['预警等级_new'] = pd.to_numeric(diff_sys_new_df['预警等级_new'], errors='coerce')
diff_sys_new_df['diff'] = diff_sys_new_df['预警等级_sys'] - diff_sys_new_df['预警等级_new']

# 将包含差异的DataFrames写入Excel文件的新工作表中
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
    diff_sys_new_df[['主体名称','预警等级_sys','预警等级_new','diff']].to_excel(writer, sheet_name='预警等级变动Sys-New', index=False)


# # -------------------------8--------------------------------------------
# # sheet 8: 预警等级每日变动详情
#
# # columns顺序示意 = ['主体名称',
# #                 '昨日预警等级_sys', '今日预警等级_sys',
# #                 '过去30天最大预警等级_sys', '过去30天最小预警等级_sys', '过去30天预警等级变动次数_sys',
# #                 '昨日预警等级_new', '今日预警等级_new',
# #                 '过去30天最大预警等级_new', '过去30天最小预警等级_new', '过去30天预警等级变动次数_new']
#
# # 假设sys_past30_df和new_past30_df是过去30天的数据
# # 初始化新的DataFrame
# comparison_df = pd.DataFrame()
#
# # 添加'主体名称'
# comparison_df['主体名称'] = sys_past30_df['主体名称'].unique()
#
# # 昨日和今日的预警等级_sys
# comparison_df = comparison_df.merge(
#     sys_past30_df[sys_past30_df['日期'] == target_date_y][['主体名称', '预警等级']],
#     how='left',
#     on='主体名称',
#     suffixes=('', '_昨日_sys')
# ).rename(columns={'预警等级': '昨日预警等级_sys'})
#
# comparison_df = comparison_df.merge(
#     sys_past30_df[sys_past30_df['日期'] == target_date][['主体名称', '预警等级']],
#     how='left',
#     on='主体名称'
# ).rename(columns={'预警等级': '今日预警等级_sys'})
#
# # 过去30天最大、最小预警等级_sys和变动次数
# sys_stats = sys_past30_df.groupby('主体名称')['预警等级'].agg(['max', 'min', pd.Series.nunique]).reset_index()
# sys_stats.columns = ['主体名称', '过去30天最大预警等级_sys', '过去30天最小预警等级_sys', '过去30天预警等级变动次数_sys']
# comparison_df = comparison_df.merge(sys_stats, on='主体名称', how='left')
#
# # 昨日和今日的预警等级_new
# comparison_df = comparison_df.merge(
#     new_past30_df[new_past30_df['日期'] == target_date_y][['主体名称', '预警等级']],
#     how='left',
#     on='主体名称',
#     suffixes=('', '_昨日_new')
# ).rename(columns={'预警等级': '昨日预警等级_new'})
#
# comparison_df = comparison_df.merge(
#     new_past30_df[new_past30_df['日期'] == target_date][['主体名称', '预警等级']],
#     how='left',
#     on='主体名称'
# ).rename(columns={'预警等级': '今日预警等级_new'})
#
# # 过去30天最大、最小预警等级_new和变动次数
# new_stats = new_past30_df.groupby('主体名称')['预警等级'].agg(['max', 'min', pd.Series.nunique]).reset_index()
# new_stats.columns = ['主体名称', '过去30天最大预警等级_new', '过去30天最小预警等级_new', '过去30天预警等级变动次数_new']
# comparison_df = comparison_df.merge(new_stats, on='主体名称', how='left')
#
# # 填充所有NaN值为0或者适当的值
# comparison_df.fillna(0, inplace=True)
#
# # 写入Excel文件
# with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
#     comparison_df.to_excel(writer, sheet_name='预警等级每日变动详情', index=False)






# -------------------------------------美化Excel-------------------------------------------
#
# 打开现有的Excel文件
workbook = load_workbook(output_path)

# -------------------------------------列宽-------------------------------------------

for sheet in workbook.sheetnames:
    worksheet = workbook[sheet]
    for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # 获取列号

            for cell in col:
                try:  # 需要尝试转换，因为值可能不是字符串
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

            adjusted_width = (max_length + 5)  # 添加小额外空间
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width


# -------------------------------------颜色-------------------------------------------
#
# index颜色
fresh_color = PatternFill(start_color='FAEBD7', end_color='FAEBD7', fill_type='solid')

for sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]
    # 设置Index行的颜色
    for cell in worksheet[1]:  # 第一行
        cell.fill = fresh_color

    # 设置Index列的颜色
    for col in worksheet.iter_cols(min_col=1, max_col=1, max_row=worksheet.max_row):  # 第一列
        for cell in col:
            cell.fill = fresh_color


# -------------------------------------分布表颜色-------------------------------------------
#
# 定义颜色：超过25%的比例时使用的亮色; 低于25%；还有其它
highlight_color = PatternFill(start_color='C497B2', end_color='C497B2', fill_type='solid') #Peru
low_color = PatternFill(start_color='F7E1ED', end_color='F7E1ED', fill_type='solid') #BrulyWood
else_color = PatternFill(start_color='F8F3F9', end_color='F8F3F9', fill_type='solid') #Bisque
# 处理的工作表名称
sheets_to_process = ["预警等级分布", "风险等级分布"]

for sheet_name in sheets_to_process:
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        # 遍历每一列，计算总和并设置颜色
        for col in worksheet.iter_cols(min_col=2, max_col=worksheet.max_column, min_row=2, max_row=worksheet.max_row):
            # 计算列的总数
            column_total = sum(cell.value for cell in col if type(cell.value) == int)

            # 如果单元格的值超过25%的列总和，设置高亮颜色；低于25%，设置低亮颜色；其余设置中等颜色。
            for cell in col:
                if type(cell.value) == int and cell.value / column_total > 0.25:
                    cell.fill = highlight_color
                elif type(cell.value) == int and cell.value / column_total > 0.1:
                    cell.fill = low_color
                else:
                    cell.fill = else_color


# -------------------------------------分布表index颜色-------------------------------------------
#
# 要填充五色的工作表
sheets_to_fill = ["预警等级分布", "风险等级分布", "风险等级变化主体透视表"]
color_mapping = {
    '高': '934B43',
    '中高': 'D76364',
    '中': 'EF7A6D',
    '低': 'F1D77E',
    '正常': 'B1CE46'
}

for sheet in sheets_to_fill:
    if sheet in workbook.sheetnames:
        ws = workbook[sheet]

        # 填充第一行
        for cell in ws[1]:
            if cell.value in risk_level_mapping:
                fill_color = color_mapping[risk_level_mapping[cell.value]]
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # 填充第一列
        for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
            for cell in row:
                if cell.value in risk_level_mapping:
                    fill_color = color_mapping[risk_level_mapping[cell.value]]
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
for sheet_name in sheets_to_fill:
    sheet = workbook[sheet_name]

    # 遍历工作表的第一行和第一列
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value in color_mapping:
                color = color_mapping[cell.value]
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    for col in sheet.iter_cols(min_col=1, max_col=1):
        for cell in col:
            if cell.value in color_mapping:
                color = color_mapping[cell.value]
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

# -------------------------------------日期注释-------------------------------------------
#
date_annotation = "Analysis based on date："+target_date  # 日期注释

# 排除不想加日期的表：
sheets_to_exclude = ["预警等级分布", "风险等级分布"]
# 循环遍历所有工作表
for sheet_name in workbook.sheetnames:
    if sheet_name not in sheets_to_exclude:
        datesheet = workbook[sheet_name]

    # 在第一行插入新的行
    datesheet.insert_rows(1)

    # 将日期注释添加到第一行的第一个单元格
    datesheet['A1'] = date_annotation

    # 扩展或格式化单元格
    for col in range(2, datesheet.max_column + 1):  # 从B列到最后一列
        datesheet[f'{get_column_letter(col)}1'] = ''


# -------------------------------------强制消除小bug-------------------------------------------
#
# 把'滑档区间'里的第一二行删掉
sheet_name = '滑档区间'

if sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]

    # 删除第一行和第二行
    worksheet.delete_rows(idx=1, amount=2)

# 保存工作簿
workbook.save(output_path)